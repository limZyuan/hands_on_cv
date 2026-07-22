"""
md_to_excel.py

Converts a raw markdown file (e.g. result.md, produced by main.py's
OCR inference) directly into an Excel workbook.

Handles TWO table formats the model can emit:
  1. Standard markdown pipe tables:   | a | b |
                                      | - | - |
                                      | 1 | 2 |
  2. Raw embedded HTML tables:        <table><tr><td colspan="2">...</td></tr></table>
     (this is what contract/form-style documents often produce)

HTML tables are parsed with BeautifulSoup, colspan is preserved as real
merged cells in Excel (not flattened/duplicated text), and HTML entities
(&#x27; &amp; etc.) plus stray LaTeX artifacts (\\( ^{th} \\)) picked up
by OCR are cleaned from the narrative text.

Usage:
    python md_to_excel.py ./outputs/result.md ./outputs/output.xlsx
"""

import os
import re
import html
import argparse
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# ── Markdown pipe-table detection ──

TABLE_ROW_RE = re.compile(r"^\s*\|(.+)\|\s*$")
TABLE_SEP_RE = re.compile(r"^\s*\|?[\s:|-]+\|?\s*$")

# ── HTML table detection ──

HTML_TABLE_RE = re.compile(r"<table\b.*?</table>", re.IGNORECASE | re.DOTALL)


def clean_text(text):
    """Unescape HTML entities and strip common OCR/LaTeX artifacts."""
    text = html.unescape(text)
    # \( ^{th} \)  ->  th      (superscript ordinal markers picked up as LaTeX)
    text = re.sub(r"\\\(\s*\^\{(\w+)\}\s*\\\)", r"\1", text)
    # Any remaining stray \( ... \) LaTeX inline-math wrappers -> keep inner content
    text = re.sub(r"\\\(\s*(.*?)\s*\\\)", r"\1", text)
    return text.strip()


def parse_html_table(table_html):
    """
    Parse one <table>...</table> block into a grid.
    Returns a list of rows, where each row is a list of
    (text, colspan) tuples — colspan is preserved so the Excel
    writer can create real merged cells instead of duplicating text.
    """
    soup = BeautifulSoup(table_html, "html.parser")
    grid = []
    for tr in soup.find_all("tr"):
        row = []
        for cell in tr.find_all(["td", "th"]):
            colspan = int(cell.get("colspan", 1) or 1)
            text = clean_text(cell.get_text(separator=" ", strip=True))
            row.append((text, colspan))
        if row:
            grid.append(row)
    return grid


def parse_markdown_pipe_tables(md_text):
    """Returns list of grids for markdown pipe tables."""
    lines = md_text.splitlines()
    results = []
    i = 0
    while i < len(lines):
        line = lines[i]
        if TABLE_ROW_RE.match(line) and i + 1 < len(lines) and TABLE_SEP_RE.match(lines[i + 1]):
            header_cells = [clean_text(c.strip()) for c in line.strip().strip("|").split("|")]
            grid = [[(c, 1) for c in header_cells]]
            i += 2
            while i < len(lines) and TABLE_ROW_RE.match(lines[i]):
                row_cells = [clean_text(c.strip()) for c in lines[i].strip().strip("|").split("|")]
                grid.append([(c, 1) for c in row_cells])
                i += 1
            results.append(grid)
            continue
        i += 1
    return results


def extract_all_tables_and_narrative(md_text):
    """
    Pulls out both HTML tables and markdown pipe tables, and returns:
      - tables: list of grids (each grid = list of rows of (text, colspan))
      - narrative_text: everything else, cleaned
    """
    tables = []

    html_grids = [parse_html_table(m.group(0)) for m in HTML_TABLE_RE.finditer(md_text)]
    text_without_html_tables = HTML_TABLE_RE.sub("\n", md_text)
    md_pipe_grids = parse_markdown_pipe_tables(text_without_html_tables)

    for grid in html_grids:
        if grid:
            tables.append(grid)
    for grid in md_pipe_grids:
        tables.append(grid)

    # Build narrative text: strip HTML tables, strip markdown pipe table lines, clean remaining text
    narrative = HTML_TABLE_RE.sub("", md_text)
    narrative_lines = narrative.splitlines()
    kept_lines = []
    j = 0
    while j < len(narrative_lines):
        line = narrative_lines[j]
        if TABLE_ROW_RE.match(line) and j + 1 < len(narrative_lines) and TABLE_SEP_RE.match(narrative_lines[j + 1]):
            j += 2
            while j < len(narrative_lines) and TABLE_ROW_RE.match(narrative_lines[j]):
                j += 1
            continue
        kept_lines.append(line)
        j += 1
    narrative_text = clean_text("\n".join(kept_lines))
    narrative_text = re.sub(r"\n{3,}", "\n\n", narrative_text).strip()

    return tables, narrative_text


# ── Excel writing ──

THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")


def sanitize_sheet_name(name, used_names):
    for ch in '[]:*?/\\':
        name = name.replace(ch, '')
    name = name[:31] or "Table"
    base = name
    i = 2
    while name in used_names:
        suffix = f"_{i}"
        name = (base[: 31 - len(suffix)]) + suffix
        i += 1
    used_names.add(name)
    return name


def grid_is_uniform_data_table(grid):
    """Heuristic: a 'real' data table has a consistent column count and no colspans."""
    col_counts = {sum(c[1] for c in row) for row in grid}
    has_colspan = any(c[1] > 1 for row in grid for c in row)
    return len(col_counts) == 1 and not has_colspan and len(grid) > 1


def write_grid_table(ws, grid, looks_like_data_table):
    """
    Write a grid (list of rows of (text, colspan)) into a worksheet,
    creating real merged cells for colspan > 1, with borders/wrapping.
    True data tables get a bolded header row; form-style (label/value)
    tables get their first column bolded as labels instead.
    """
    max_cols = max(sum(c[1] for c in row) for row in grid)

    for r_idx, row in enumerate(grid, start=1):
        col_idx = 1
        for text, colspan in row:
            cell = ws.cell(row=r_idx, column=col_idx, value=text if text else None)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = Font(name="Arial")
            if colspan > 1:
                end_col = min(col_idx + colspan - 1, max_cols)
                if end_col > col_idx:
                    ws.merge_cells(start_row=r_idx, start_column=col_idx, end_row=r_idx, end_column=end_col)
            col_idx += colspan

    if looks_like_data_table:
        for cell in ws[1]:
            cell.font = Font(bold=True, name="Arial")
            cell.fill = HEADER_FILL
    else:
        for row in ws.iter_rows():
            first_cell = row[0]
            if first_cell.value:
                first_cell.font = Font(bold=True, name="Arial")
                first_cell.fill = HEADER_FILL

    for col_idx in range(1, max_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 10
        for row in grid:
            running = 0
            for text, colspan in row:
                running += colspan
                if running >= col_idx > running - colspan:
                    max_len = max(max_len, len(text))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 60)


def write_excel(md_path, output_path):
    with open(md_path, "r", encoding="utf-8") as f:
        md_text = f.read()

    tables, narrative_text = extract_all_tables_and_narrative(md_text)
    source_file = os.path.basename(md_path)

    wb = Workbook()
    wb.remove(wb.active)
    used_names = set()

    if tables:
        for grid in tables:
            sheet_name = sanitize_sheet_name(f"Table_{len(used_names) + 1}", used_names)
            ws = wb.create_sheet(sheet_name)
            write_grid_table(ws, grid, grid_is_uniform_data_table(grid))
    else:
        ws = wb.create_sheet(sanitize_sheet_name("Table_1", used_names))
        ws["A1"] = "No tables were detected in the source document."

    if narrative_text:
        ws_n = wb.create_sheet(sanitize_sheet_name("Narrative", used_names))
        ws_n["A1"] = "Contract Text"
        ws_n["A1"].font = Font(bold=True, name="Arial")
        for i, line in enumerate(narrative_text.split("\n"), start=2):
            cell = ws_n.cell(row=i, column=1, value=line if line else None)
            cell.font = Font(name="Arial")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws_n.column_dimensions["A"].width = 100

    ws_m = wb.create_sheet(sanitize_sheet_name("Metadata", used_names))
    ws_m["A1"], ws_m["B1"] = "Field", "Value"
    ws_m["A1"].font = Font(bold=True, name="Arial")
    ws_m["B1"].font = Font(bold=True, name="Arial")
    ws_m["A2"], ws_m["B2"] = "Source File", source_file
    ws_m["A2"].font = Font(name="Arial")
    ws_m["B2"].font = Font(name="Arial")

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Convert a raw markdown file (with embedded HTML tables) into a formatted Excel workbook.")
    parser.add_argument("md_path", help="Path to the markdown file (e.g. result.md)")
    parser.add_argument("output_xlsx", help="Path to write the .xlsx file")
    args = parser.parse_args()

    write_excel(args.md_path, args.output_xlsx)
    print(f"Wrote {args.output_xlsx}")


if __name__ == "__main__":
    main()
