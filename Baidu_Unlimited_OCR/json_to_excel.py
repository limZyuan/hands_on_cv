"""
json_to_excel.py

Companion script to main.py.
Reads structured.json (produced by main.py's build_structured_output)
and writes an .xlsx workbook: one sheet per table found in the contract,
plus a "Narrative" sheet holding any non-tabular text (clauses, terms, etc.).

structured.json schema (as produced by main.py):
{
  "source_file": "...",
  "tables": [
    {
      "table_index": 1,
      "grid": [ [[text, colspan], [text, colspan], ...], ... ],
      "is_data_table": true/false
    },
    ...
  ],
  "narrative_text": "..."
}

Colspan is preserved from the original HTML/markdown tables, so cells
that were merged in the source document are written as real merged
Excel cells here — not duplicated or flattened text.

Usage:
    python json_to_excel.py ./outputs/structured.json ./outputs/output.xlsx
"""

import json
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")


def sanitize_sheet_name(name, used_names):
    """Excel sheet names: max 31 chars, no []:*?/\\, must be unique."""
    for ch in '[]:*?/\\':
        name = name.replace(ch, '')
    name = name[:31] or "Table"
    base = name
    i = 2
    while name in used_names:
        suffix = f"_{i}"
        name = (base[: 31 - len(suffix)]) + suffix
        i += 1
    used_names.add(name)
    return name


def write_grid_table(ws, grid, is_data_table):
    """
    Write a grid (list of rows of [text, colspan]) into a worksheet,
    creating real merged cells for colspan > 1, with borders/wrapping.
    True data tables get a bolded header row; form-style (label/value)
    tables get their first column bolded as labels instead.
    """
    max_cols = max(sum(c[1] for c in row) for row in grid)

    for r_idx, row in enumerate(grid, start=1):
        col_idx = 1
        for text, colspan in row:
            cell = ws.cell(row=r_idx, column=col_idx, value=text if text else None)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = Font(name="Arial")
            if colspan > 1:
                end_col = min(col_idx + colspan - 1, max_cols)
                if end_col > col_idx:
                    ws.merge_cells(start_row=r_idx, start_column=col_idx, end_row=r_idx, end_column=end_col)
            col_idx += colspan

    if is_data_table:
        for cell in ws[1]:
            cell.font = Font(bold=True, name="Arial")
            cell.fill = HEADER_FILL
    else:
        # Form-style table: bold the first column (labels)
        for row in ws.iter_rows():
            first_cell = row[0]
            if first_cell.value:
                first_cell.font = Font(bold=True, name="Arial")
                first_cell.fill = HEADER_FILL

    for col_idx in range(1, max_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 10
        for row in grid:
            running = 0
            for text, colspan in row:
                running += colspan
                if running >= col_idx > running - colspan:
                    max_len = max(max_len, len(text))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 60)


def write_excel(structured, output_path):
    tables = structured.get("tables", [])
    narrative = structured.get("narrative_text", "")
    source_file = structured.get("source_file", "")

    wb = Workbook()
    wb.remove(wb.active)
    used_names = set()

    if tables:
        for t in tables:
            sheet_name = sanitize_sheet_name(f"Table_{t['table_index']}", used_names)
            ws = wb.create_sheet(sheet_name)
            write_grid_table(ws, t["grid"], t.get("is_data_table", False))
    else:
        ws = wb.create_sheet(sanitize_sheet_name("Table_1", used_names))
        ws["A1"] = "No tables were detected in the source document."

    if narrative.strip():
        ws_n = wb.create_sheet(sanitize_sheet_name("Narrative", used_names))
        ws_n["A1"] = "Contract Text"
        ws_n["A1"].font = Font(bold=True, name="Arial")
        for i, line in enumerate(narrative.split("\n"), start=2):
            cell = ws_n.cell(row=i, column=1, value=line if line else None)
            cell.font = Font(name="Arial")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws_n.column_dimensions["A"].width = 100

    ws_m = wb.create_sheet(sanitize_sheet_name("Metadata", used_names))
    ws_m["A1"], ws_m["B1"] = "Field", "Value"
    ws_m["A1"].font = Font(bold=True, name="Arial")
    ws_m["B1"].font = Font(bold=True, name="Arial")
    ws_m["A2"], ws_m["B2"] = "Source File", source_file
    ws_m["A2"].font = Font(name="Arial")
    ws_m["B2"].font = Font(name="Arial")

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="Convert structured.json (from main.py) into a formatted Excel workbook.")
    parser.add_argument("json_path", help="Path to structured.json")
    parser.add_argument("output_xlsx", help="Path to write the .xlsx file")
    args = parser.parse_args()

    with open(args.json_path, "r", encoding="utf-8") as f:
        structured = json.load(f)

    write_excel(structured, args.output_xlsx)
    print(f"Wrote {args.output_xlsx}")


if __name__ == "__main__":
    main()
